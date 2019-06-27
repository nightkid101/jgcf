from openpyxl import Workbook, load_workbook
from datetime import datetime
from pymongo import MongoClient
from init import config_init

config = config_init()

if config['mongodb']['dev_mongo'] == '1':
    db = MongoClient(config['mongodb']['ali_mongodb_url'], username=config['mongodb']['ali_mongodb_username'],
                     password=config['mongodb']['ali_mongodb_password'],
                     port=int(config['mongodb']['ali_mongodb_port']))[config['mongodb']['ali_mongodb_name']]
else:
    db = MongoClient(
        host=config['mongodb']['mongodb_host'],
        port=int(config['mongodb']['mongodb_port']),
        username=None if config['mongodb']['mongodb_username'] == '' else config['mongodb']['mongodb_username'],
        password=None if config['mongodb']['mongodb_password'] == '' else config['mongodb']['mongodb_password'])[
        config['mongodb']['mongodb_db_name']]

company_keywords_list = ['A股证券代码', 'A股证券简称', 'B股证券代码', 'B股证券简称', '新三板证券代码', '新三板证券简称',
                         '公司名称', '涉及公司', '职务/角色', '注册地址', '住所地址', '办公地址', '法定代表人',
                         '负责人', '一码通代码', '工商注册号', '注册号', '机构代码', '登记时间', '涉及对象', '法人代表证件号', '成立日期']
person_keywords_list = ['姓名', '性别', '年龄', '出生日期', '国籍', '民族', '住所地址', '身份证号', '任职期间', '就职公司', '职务/角色',
                        '执业类别', '注册号', '资格证号', '执业证号', '登记编号', '一码通代码', '登记时间', '取得证书时间', '人物关系',
                        '港澳台证件号码', '护照号', '办公地址']


def litigant_xlsx(org):
    file = Workbook()
    sheet = file.active
    sheet.title = "当事人解析结果"
    sheet.append(['id', 'url', '当事人', '解析结果字段名', '解析结果字段值'])
    count = 2
    for each_litigant_parsed_result in db.litigant_parsed_result.find({'org': org}):
        sheet['A' + str(count)] = str(each_litigant_parsed_result['_id'])
        sheet['B' + str(count)] = str(each_litigant_parsed_result['origin_litigant'])
        try:
            sheet['C' + str(count)] = db.parsed_data.find_one(
                {'_id': db.announcement.find_one({'_id': each_litigant_parsed_result['origin_announcement_id']})[
                    'oss_file_id']})['oss_file_origin_url']
        except Exception as e:
            sheet['C' + str(count)] = ''
        for each_litigant_info in each_litigant_parsed_result['parsed_result']:
            each_info = {}
            if each_litigant_info.get('公司名称', '') != '':
                for each_keyword in company_keywords_list:
                    if each_litigant_info.get(each_keyword, '') != '':
                        each_info[each_keyword] = each_litigant_info.get(each_keyword, '')
                    else:
                        each_info[each_keyword] = ''
                for each_key in each_info:
                    sheet['D' + str(count)] = str(each_key)
                    sheet['E' + str(count)] = str(each_info[each_key])
                    count += 1
                count += 1
            else:
                if each_litigant_info.get('姓名', '') != '':
                    for each_keyword in person_keywords_list:
                        if each_litigant_info.get(each_keyword, '') != '':
                            each_info[each_keyword] = each_litigant_info.get(each_keyword, '')
                        else:
                            each_info[each_keyword] = ''
                    for each_key in each_info:
                        sheet['D' + str(count)] = str(each_key)
                        sheet['E' + str(count)] = str(each_info[each_key])
                        count += 1
                    count += 1
        count += 1

    today_date = datetime.now().date()
    file.save('/Users/austinzy/Desktop/' + '当事人解析结果' + org + today_date.strftime("%Y年%m月%d日") + '.xlsx')


def abb_full_xlsx():
    file = Workbook()
    sheet = file.active
    sheet.title = "全称简称对应"
    sheet.append(['全称', '简称', 'url'])
    count = 2
    for each_litigant_abb_full_result in db.litigant_abb_full_result.find(no_cursor_timeout=True):
        sheet['A' + str(count)] = str(each_litigant_abb_full_result['fullName'])
        sheet['B' + str(count)] = str(each_litigant_abb_full_result['abbreviation'])
        sheet['C' + str(count)] = str(each_litigant_abb_full_result['url'])
        count += 1
    file.save('/Users/austinzy/Desktop/' + '全称简称对应1117' + '.xlsx')


litigant_xlsx('银监机构')
abb_full_xlsx()
